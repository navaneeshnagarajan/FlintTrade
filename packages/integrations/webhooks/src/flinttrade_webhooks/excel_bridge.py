"""Bridge for Excel ↔ FlintTrade data exchange.

Supports:
  - Export positions/orders/holdings/any list[dict] to Excel (.xlsx)
  - Import watchlists or any tabular data from Excel
  - Create a formatted multi-sheet portfolio report

openpyxl is an optional dependency. If it is not installed, ``ExcelBridgeError``
is raised at call time with a clear installation message — the module still
imports successfully so other integration features are unaffected.

Install openpyxl::

    pip install openpyxl

Usage::

    bridge = ExcelBridge()
    path = bridge.export_to_excel(positions, "Positions", "/tmp/positions.xlsx")
    rows = bridge.import_from_excel("/tmp/watchlist.xlsx", sheet_name="Watchlist")
    report = bridge.create_portfolio_report(positions, holdings, "/tmp/portfolio.xlsx")
"""

from __future__ import annotations

import logging
import os
from typing import Any

logger = logging.getLogger("flinttrade.integration.excel_bridge")

# ---------------------------------------------------------------------------
# Optional dependency: openpyxl
# ---------------------------------------------------------------------------

try:
    import openpyxl  # noqa: F401
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

_OPENPYXL_MISSING_MSG = (
    "openpyxl is not installed. Install it with: pip install openpyxl"
)

# Header style constants
_HEADER_BG = "1E3A5F"   # Dark blue matching FlintTrade palette
_HEADER_FG = "FFFFFF"   # White text
_POSITIVE_FG = "22C55E" # Green for profits
_NEGATIVE_FG = "EF4444" # Red for losses


class ExcelBridgeError(Exception):
    """Raised when an Excel operation fails."""


def _require_openpyxl() -> None:
    """Raise a clear error if openpyxl is not installed."""
    if not _OPENPYXL_AVAILABLE:
        raise ExcelBridgeError(_OPENPYXL_MISSING_MSG)


class ExcelBridge:
    """Bridge for Excel ↔ FlintTrade data exchange.

    All methods require openpyxl. If not installed, ``ExcelBridgeError`` is
    raised with a pip install instruction.

    Example::

        bridge = ExcelBridge()
        bridge.export_to_excel(positions, "Positions", "~/positions.xlsx")
        bridge.create_portfolio_report(positions, holdings, "~/portfolio.xlsx")
    """

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_to_excel(
        self,
        data: list[dict[str, Any]],
        sheet_name: str,
        file_path: str,
    ) -> str:
        """Export a list of dicts to an Excel file.

        Creates a new workbook with a single sheet. Column headers are derived
        from the keys of the first dict. All columns are auto-sized.

        Args:
            data: List of row dicts. All dicts should share the same keys.
            sheet_name: Worksheet name (max 31 chars, Excel limit).
            file_path: Output ``.xlsx`` file path. Parent directory must exist.

        Returns:
            Absolute path to the written file.

        Raises:
            ExcelBridgeError: If openpyxl is not installed or file write fails.
        """
        _require_openpyxl()
        wb = self._build_data_workbook(data, sheet_name)
        self._save(wb, file_path)
        logger.info("Exported %d rows to %s (sheet=%s)", len(data), file_path, sheet_name)
        return os.path.abspath(file_path)

    def _build_data_workbook(self, data: list[dict[str, Any]], sheet_name: str) -> Any:
        """Build a single-sheet workbook from ``data`` (shared by file + bytes export)."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel worksheet name limit

        if data:
            headers = list(data[0].keys())
            self._write_header_row(ws, headers)
            for row_dict in data:
                ws.append([row_dict.get(h) for h in headers])
            self._autosize_columns(ws)
        return wb

    def export_to_bytes(
        self,
        data: list[dict[str, Any]],
        sheet_name: str = "Data",
    ) -> bytes:
        """Build the same workbook as :meth:`export_to_excel` but return the raw
        ``.xlsx`` bytes instead of writing a server-side file.

        This is the streaming path: the terminal POSTs rows and the backend
        streams the workbook straight back as a browser download, so the file
        never has to live on the server's disk.

        Args:
            data: List of row dicts (all should share the same keys).
            sheet_name: Worksheet name (max 31 chars, Excel limit).

        Returns:
            The ``.xlsx`` file content as bytes.

        Raises:
            ExcelBridgeError: If openpyxl is not installed or the build fails.
        """
        _require_openpyxl()
        from io import BytesIO  # noqa: PLC0415

        wb = self._build_data_workbook(data, sheet_name)
        buf = BytesIO()
        try:
            wb.save(buf)
        except Exception as exc:
            raise ExcelBridgeError(f"Failed to build Excel workbook: {exc}") from exc
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    def import_from_excel(
        self,
        file_path: str,
        sheet_name: str = "Sheet1",
    ) -> list[dict[str, Any]]:
        """Import data from an Excel file.

        Reads the first row as column headers and converts each subsequent
        row into a dict. Empty rows are skipped.

        Args:
            file_path: Path to the ``.xlsx`` file.
            sheet_name: Worksheet name to read. Default ``"Sheet1"``.

        Returns:
            List of row dicts where keys are the column headers.

        Raises:
            ExcelBridgeError: If openpyxl is not installed, the file does not
                exist, or the sheet is not found.
        """
        _require_openpyxl()

        if not os.path.exists(file_path):
            raise ExcelBridgeError(f"File not found: {file_path}")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelBridgeError(f"Failed to open {file_path}: {exc}") from exc

        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ExcelBridgeError(
                f"Sheet '{sheet_name}' not found in {file_path}. "
                f"Available sheets: {available}"
            )

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

        result: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            if all(cell is None for cell in raw_row):
                continue  # skip blank rows
            result.append(dict(zip(headers, raw_row)))

        logger.info("Imported %d rows from %s (sheet=%s)", len(result), file_path, sheet_name)
        return result

    # ------------------------------------------------------------------
    # Portfolio report
    # ------------------------------------------------------------------

    def create_portfolio_report(
        self,
        positions: list[dict[str, Any]],
        holdings: list[dict[str, Any]],
        file_path: str,
    ) -> str:
        """Create a formatted multi-sheet portfolio report in Excel.

        Sheet layout:
          - **Positions** — intraday/F&O positions with P&L colour-coding
          - **Holdings** — equity holdings with unrealised P&L
          - **Summary** — aggregated totals (realised P&L, total invested, etc.)

        Args:
            positions: List of position dicts (from OpenAlgo ``positionbook``).
            holdings: List of holding dicts (from OpenAlgo ``holdings``).
            file_path: Output ``.xlsx`` file path.

        Returns:
            Absolute path to the written file.

        Raises:
            ExcelBridgeError: If openpyxl is not installed or file write fails.
        """
        _require_openpyxl()
        wb = self._build_portfolio_workbook(positions, holdings)
        self._save(wb, file_path)
        logger.info(
            "Portfolio report written: %s (%d positions, %d holdings)",
            file_path, len(positions), len(holdings),
        )
        return os.path.abspath(file_path)

    def _build_portfolio_workbook(
        self,
        positions: list[dict[str, Any]],
        holdings: list[dict[str, Any]],
    ) -> Any:
        """Build the 3-sheet portfolio workbook (shared by file + bytes export)."""
        wb = Workbook()

        # -- Positions sheet --
        ws_pos = wb.active
        ws_pos.title = "Positions"
        self._write_sheet(ws_pos, positions, pnl_field="pnl")

        # -- Holdings sheet --
        ws_hold = wb.create_sheet("Holdings")
        self._write_sheet(ws_hold, holdings, pnl_field="pnl")

        # -- Summary sheet --
        ws_sum = wb.create_sheet("Summary")
        self._write_summary(ws_sum, positions, holdings)
        return wb

    def create_portfolio_report_bytes(
        self,
        positions: list[dict[str, Any]],
        holdings: list[dict[str, Any]],
    ) -> bytes:
        """Build the same 3-sheet portfolio report as
        :meth:`create_portfolio_report` but return the ``.xlsx`` bytes for
        streaming a browser download (no server-side file).

        Raises:
            ExcelBridgeError: If openpyxl is not installed or the build fails.
        """
        _require_openpyxl()
        from io import BytesIO  # noqa: PLC0415

        wb = self._build_portfolio_workbook(positions, holdings)
        buf = BytesIO()
        try:
            wb.save(buf)
        except Exception as exc:
            raise ExcelBridgeError(f"Failed to build portfolio report: {exc}") from exc
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_header_row(ws: Any, headers: list[str]) -> None:
        """Write a styled header row to the worksheet."""
        ws.append(headers)
        header_row = ws[1]
        for cell in header_row:
            cell.font = Font(bold=True, color=_HEADER_FG, name="Calibri", size=11)
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=_HEADER_BG,
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    @staticmethod
    def _autosize_columns(ws: Any, min_width: int = 10, max_width: int = 40) -> None:
        """Auto-size columns based on the maximum content length."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 2, min_width), max_width
            )

    def _write_sheet(
        self,
        ws: Any,
        data: list[dict[str, Any]],
        pnl_field: str = "pnl",
    ) -> None:
        """Write data to a worksheet with header row and P&L colour-coding."""
        if not data:
            ws.append(["No data available"])
            return

        headers = list(data[0].keys())
        self._write_header_row(ws, headers)

        pnl_col_idx: int | None = None
        if pnl_field in headers:
            pnl_col_idx = headers.index(pnl_field) + 1  # 1-indexed

        for row_num, row_dict in enumerate(data, start=2):
            row_vals = [row_dict.get(h) for h in headers]
            ws.append(row_vals)

            # Colour P&L cells
            if pnl_col_idx is not None:
                pnl_val = row_dict.get(pnl_field)
                if pnl_val is not None:
                    try:
                        cell = ws.cell(row=row_num, column=pnl_col_idx)
                        colour = _POSITIVE_FG if float(pnl_val) >= 0 else _NEGATIVE_FG
                        cell.font = Font(bold=True, color=colour)
                    except (ValueError, TypeError):
                        pass

        self._autosize_columns(ws)

    @staticmethod
    def _write_summary(
        ws: Any,
        positions: list[dict[str, Any]],
        holdings: list[dict[str, Any]],
    ) -> None:
        """Write a summary sheet with aggregated totals."""
        ws.title = "Summary"

        def _sum_field(rows: list[dict[str, Any]], field: str) -> float:
            total = 0.0
            for row in rows:
                try:
                    total += float(row.get(field) or 0)
                except (ValueError, TypeError):
                    pass
            return total

        realised_pnl = _sum_field(positions, "pnl")
        unrealised_pnl = _sum_field(holdings, "pnl")
        total_invested = _sum_field(holdings, "buy_avg") * len(holdings)

        summary_rows = [
            ("Metric", "Value"),
            ("Total Positions", len(positions)),
            ("Total Holdings", len(holdings)),
            ("Realised P&L (Positions)", round(realised_pnl, 2)),
            ("Unrealised P&L (Holdings)", round(unrealised_pnl, 2)),
            ("Net P&L", round(realised_pnl + unrealised_pnl, 2)),
            ("Total Invested (est.)", round(total_invested, 2)),
        ]

        for row in summary_rows:
            ws.append(list(row))

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color=_HEADER_FG, name="Calibri", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center")

        # Colour net P&L
        net_pnl = realised_pnl + unrealised_pnl
        net_cell = ws.cell(row=len(summary_rows), column=2)
        net_cell.font = Font(
            bold=True,
            color=_POSITIVE_FG if net_pnl >= 0 else _NEGATIVE_FG,
        )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _save(wb: Any, file_path: str) -> None:
        """Save workbook to file, raising ExcelBridgeError on failure."""
        try:
            wb.save(file_path)
        except Exception as exc:
            raise ExcelBridgeError(f"Failed to save Excel file {file_path}: {exc}") from exc
