"""Tests for ExcelBridge — openpyxl write/read round-trip, graceful fallback.

Tests that require real file I/O use a temporary directory.
Tests for the graceful fallback mock out _OPENPYXL_AVAILABLE.
"""

from __future__ import annotations

import os
import tempfile


class TestExcelBridgeExportImport:
    """export_to_excel + import_from_excel round-trip (requires openpyxl)."""

    def _skip_if_no_openpyxl(self):
        """Skip silently if openpyxl is not installed."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            import pytest
            pytest.skip("openpyxl not installed")

    def test_export_and_reimport_round_trip(self):
        """Exported data must survive an import round-trip without loss."""
        self._skip_if_no_openpyxl()
        from flinttrade_webhooks.excel_bridge import ExcelBridge

        data = [
            {"symbol": "RELIANCE", "exchange": "NSE", "qty": 10, "avg_price": 2980.5},
            {"symbol": "TCS", "exchange": "NSE", "qty": 5, "avg_price": 3520.0},
        ]

        bridge = ExcelBridge()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_export.xlsx")
            returned_path = bridge.export_to_excel(data, "Holdings", path)
            assert os.path.exists(returned_path)

            imported = bridge.import_from_excel(returned_path, "Holdings")

        assert len(imported) == 2
        assert imported[0]["symbol"] == "RELIANCE"
        assert float(imported[0]["avg_price"]) == 2980.5
        assert imported[1]["symbol"] == "TCS"

    def test_export_empty_data_creates_valid_file(self):
        """export_to_excel with empty list should still create a valid file."""
        self._skip_if_no_openpyxl()
        from flinttrade_webhooks.excel_bridge import ExcelBridge

        bridge = ExcelBridge()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "empty.xlsx")
            returned_path = bridge.export_to_excel([], "EmptySheet", path)
            assert os.path.exists(returned_path)

    def test_export_to_bytes_round_trips_via_openpyxl(self):
        """export_to_bytes must yield a valid .xlsx that re-imports losslessly."""
        self._skip_if_no_openpyxl()
        import io

        from openpyxl import load_workbook

        from flinttrade_webhooks.excel_bridge import ExcelBridge

        data = [
            {"symbol": "RELIANCE", "qty": 10, "pnl": 1200.5},
            {"symbol": "TCS", "qty": 5, "pnl": -300.0},
        ]
        bridge = ExcelBridge()
        payload = bridge.export_to_bytes(data, "Positions")

        assert isinstance(payload, bytes)
        # Valid .xlsx files are ZIP archives → start with the PK magic bytes.
        assert payload[:2] == b"PK"

        wb = load_workbook(io.BytesIO(payload))
        ws = wb["Positions"]
        # Header row + 2 data rows.
        assert ws.max_row == 3
        assert ws.cell(row=1, column=1).value == "symbol"
        assert ws.cell(row=2, column=1).value == "RELIANCE"

    def test_export_to_bytes_empty_data_is_valid(self):
        """export_to_bytes with no rows still produces a valid workbook."""
        self._skip_if_no_openpyxl()
        import io

        from openpyxl import load_workbook

        from flinttrade_webhooks.excel_bridge import ExcelBridge

        payload = ExcelBridge().export_to_bytes([], "Empty")
        assert payload[:2] == b"PK"
        wb = load_workbook(io.BytesIO(payload))
        assert "Empty" in wb.sheetnames

    def test_import_raises_on_missing_file(self):
        """import_from_excel raises ExcelBridgeError if file does not exist."""
        self._skip_if_no_openpyxl()
        from flinttrade_webhooks.excel_bridge import ExcelBridge, ExcelBridgeError

        bridge = ExcelBridge()
        try:
            bridge.import_from_excel("/nonexistent/path/file.xlsx")
            assert False, "Expected ExcelBridgeError"
        except ExcelBridgeError as exc:
            assert "not found" in str(exc).lower()

    def test_import_raises_on_missing_sheet(self):
        """import_from_excel raises ExcelBridgeError for a non-existent sheet."""
        self._skip_if_no_openpyxl()
        from flinttrade_webhooks.excel_bridge import ExcelBridge, ExcelBridgeError

        data = [{"col": "val"}]
        bridge = ExcelBridge()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test.xlsx")
            bridge.export_to_excel(data, "Sheet1", path)
            try:
                bridge.import_from_excel(path, sheet_name="NonExistentSheet")
                assert False, "Expected ExcelBridgeError"
            except ExcelBridgeError as exc:
                assert "NonExistentSheet" in str(exc)


class TestExcelBridgePortfolioReport:
    """create_portfolio_report — multi-sheet report with summary."""

    def _skip_if_no_openpyxl(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            import pytest
            pytest.skip("openpyxl not installed")

    def test_portfolio_report_creates_three_sheets(self):
        """Report must have Positions, Holdings, and Summary sheets."""
        self._skip_if_no_openpyxl()
        from openpyxl import load_workbook

        from flinttrade_webhooks.excel_bridge import ExcelBridge

        positions = [
            {"symbol": "NIFTY26APR24500CE", "exchange": "NFO", "qty": 75, "pnl": 1200.0},
        ]
        holdings = [
            {"symbol": "RELIANCE", "exchange": "NSE", "qty": 10, "buy_avg": 2900.0, "pnl": 800.0},
        ]

        bridge = ExcelBridge()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "portfolio.xlsx")
            returned_path = bridge.create_portfolio_report(positions, holdings, path)
            assert os.path.exists(returned_path)

            wb = load_workbook(returned_path)
            assert "Positions" in wb.sheetnames
            assert "Holdings" in wb.sheetnames
            assert "Summary" in wb.sheetnames

    def test_portfolio_report_bytes_round_trips(self):
        """create_portfolio_report_bytes yields a valid multi-sheet .xlsx."""
        self._skip_if_no_openpyxl()
        import io

        from openpyxl import load_workbook

        from flinttrade_webhooks.excel_bridge import ExcelBridge

        positions = [{"symbol": "NIFTY26APR24500CE", "qty": 75, "pnl": 1200.0}]
        holdings = [{"symbol": "RELIANCE", "qty": 10, "buy_avg": 2900.0, "pnl": 800.0}]

        payload = ExcelBridge().create_portfolio_report_bytes(positions, holdings)
        assert isinstance(payload, bytes)
        assert payload[:2] == b"PK"  # valid .xlsx (zip) magic

        wb = load_workbook(io.BytesIO(payload))
        assert {"Positions", "Holdings", "Summary"}.issubset(set(wb.sheetnames))


class TestExcelBridgeGracefulFallback:
    """Graceful fallback when openpyxl is not installed."""

    def test_export_raises_excel_bridge_error_without_openpyxl(self):
        """All methods should raise ExcelBridgeError when openpyxl is absent."""
        import flinttrade_webhooks.excel_bridge as eb_module
        from flinttrade_webhooks.excel_bridge import ExcelBridge, ExcelBridgeError

        original = eb_module._OPENPYXL_AVAILABLE
        try:
            eb_module._OPENPYXL_AVAILABLE = False
            bridge = ExcelBridge()
            try:
                bridge.export_to_excel([], "Sheet1", "/tmp/x.xlsx")
                assert False, "Expected ExcelBridgeError"
            except ExcelBridgeError as exc:
                assert "openpyxl" in str(exc).lower()
        finally:
            eb_module._OPENPYXL_AVAILABLE = original

    def test_import_raises_excel_bridge_error_without_openpyxl(self):
        import flinttrade_webhooks.excel_bridge as eb_module
        from flinttrade_webhooks.excel_bridge import ExcelBridge, ExcelBridgeError

        original = eb_module._OPENPYXL_AVAILABLE
        try:
            eb_module._OPENPYXL_AVAILABLE = False
            bridge = ExcelBridge()
            try:
                bridge.import_from_excel("/tmp/x.xlsx")
                assert False, "Expected ExcelBridgeError"
            except ExcelBridgeError as exc:
                assert "openpyxl" in str(exc).lower()
        finally:
            eb_module._OPENPYXL_AVAILABLE = original

    def test_module_imports_cleanly_without_openpyxl(self):
        """The module must import successfully even when openpyxl is absent."""
        # If we reach here, the import succeeded (tested implicitly by any import above).
        import flinttrade_webhooks.excel_bridge as eb_module  # noqa: F401
        assert hasattr(eb_module, "ExcelBridge")
        assert hasattr(eb_module, "ExcelBridgeError")
